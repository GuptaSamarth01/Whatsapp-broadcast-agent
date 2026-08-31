"""
Excel Contact Deduplication Utility
------------------------------------
Reads phone numbers from Sheet2 of an Excel workbook,
removes duplicates, and saves the unique numbers as JSON.
"""

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize_phone(phone) -> str | None:
    """Normalize a phone number for duplicate detection."""

    if phone is None:
        return None

    phone = str(phone).strip()

    if not phone:
        return None

    # Remove spaces, brackets, hyphens, etc.
    phone = re.sub(r"[^\d+]", "", phone)

    # Convert 00XXXXXXXXXX → +XXXXXXXXXX
    if phone.startswith("00"):
        phone = "+" + phone[2:]

    # Add + if missing.
    if not phone.startswith("+"):
        phone = "+" + phone

    return phone


def remove_duplicate_numbers(
    input_path: str,
    output_path: str,
    sheet_name: str = "Sheet2",
) -> None:
    """Extract unique phone numbers and save them as JSON."""

    input_file = Path(input_path)
    output_file = Path(output_path)

    if not input_file.exists():
        raise FileNotFoundError(
            f"Input file not found: {input_file}"
        )

    workbook = load_workbook(input_file)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found."
        )

    sheet = workbook[sheet_name]

    seen = set()
    unique_numbers = []

    original_count = 0

    # Row 1 is assumed to contain the header.
    for row in range(2, sheet.max_row + 1):

        value = sheet.cell(
            row=row,
            column=1,
        ).value

        if value is None:
            continue

        phone = normalize_phone(value)

        if phone is None:
            continue

        original_count += 1

        if phone not in seen:

            seen.add(phone)
            unique_numbers.append(phone)

    # Save unique numbers as JSON.
    with open(
        output_file,
        "w",
        encoding="utf-8",
    ) as f:

        json.dump(
            unique_numbers,
            f,
            indent=2,
            ensure_ascii=False,
        )

    print("\n===================================")
    print("      DUPLICATE REMOVAL COMPLETE")
    print("===================================\n")

    print(f"Input:              {input_file}")
    print(f"Sheet:              {sheet_name}")
    print(f"Original numbers:   {original_count}")
    print(f"Unique numbers:     {len(unique_numbers)}")
    print(
        f"Duplicates removed: "
        f"{original_count - len(unique_numbers)}"
    )
    print(f"\nJSON output:        {output_file}")

    print("\n===================================\n")


def main() -> None:
    """Command-line interface."""

    if len(sys.argv) < 3:

        print("\nUsage:")
        print(
            "  python remove_duplicates.py "
            "input.xlsx output.json"
        )

        print("\nExample:")
        print(
            "  python remove_duplicates.py "
            "contacts.xlsx contacts.json"
        )

        return

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    try:

        remove_duplicate_numbers(
            input_path=input_path,
            output_path=output_path,
        )

    except Exception as e:

        print(f"\nError: {e}")

        sys.exit(1)


if __name__ == "__main__":
    main()
